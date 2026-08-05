"""Build the DRE (Demonstração do Resultado do Exercício) sheet for the Fechamento report.

The DRE is added as a new worksheet in the workbook produced by ``build_report.py``.
Values derived from NF-e data reference the Detalhado sheet via live Excel formulas;
lines without source data default to zero and can be filled manually.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Literal

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from report_common import (
    BORDER,
    LABEL_FONT,
    MONEY_FMT,
    NOTE_FONT,
    SECTION_FILL,
    SRC_DESCONTO,
    SRC_FATURADO,
    SRC_FRETE,
    TITLE_FONT,
    TOTAL_FILL,
    as_excel_label,
    style_header_cell,
    sum_cancelled_sales,
    sum_column,
    sum_columns,
    sum_return_cfops,
    sumifs_wildcard,
)

RowKind = Literal["section", "item", "subtotal", "total"]


def _row_ref(row: int, col: int = 2) -> str:
    return f"{get_column_letter(col)}{row}"


def build_dre(
    ws: Worksheet,
    det_col: dict[str, str],
    last_row: int,
    *,
    sales_tax_rate_cell: str = "Parametros!$G$3",
) -> None:
    """Write the DRE sheet (mutates *ws* in place)."""
    ws["A1"] = "DRE - Demonstração do Resultado do Exercício"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    style_header_cell(ws.cell(row=header_row, column=1, value="Conta"))
    style_header_cell(ws.cell(row=header_row, column=2, value="Valor (R$)"))

    sections: list[tuple[str, list[tuple[str, RowKind, str | None, int]]]] = [
        (
            "RECEITA OPERACIONAL BRUTA",
            [
                ("Faturamento Bruto", "item", "faturamento_bruto", 1),
                ("(-) Vendas Canceladas", "item", "vendas_canceladas", 1),
                ("(-) Devoluções/Reembolsos", "item", "devolucoes", 1),
                ("(-) Descontos Concedidos", "item", "descontos", 1),
                ("= Receita Bruta Ajustada", "subtotal", "receita_bruta_ajustada", 0),
            ],
        ),
        (
            "DEDUÇÕES DA RECEITA",
            [
                ("(-) Impostos sobre Vendas (4%)", "item", "impostos_vendas", 1),
                ("= RECEITA LÍQUIDA", "subtotal", "receita_liquida", 0),
            ],
        ),
        (
            "CUSTOS VARIÁVEIS",
            [
                ("(-) CMV (Custo das Mercadorias Vendidas)", "item", None, 1),
                ("(-) Taxas de Marketplace", "item", "taxas_marketplace", 1),
                ("(-) Taxas de Envio/Frete Pago", "item", "frete", 1),
                ("(-) Custo com Expedição", "item", None, 1),
                ("(-) Embalagens", "item", None, 1),
                ("= LUCRO BRUTO", "subtotal", "lucro_bruto", 0),
            ],
        ),
        (
            "DESPESAS COMERCIAL / MARKETING",
            [
                ("(-) ADS Mercado Livre", "item", "ads_ml", 1),
                ("(-) ADS Shopee", "item", "ads_shopee", 1),
                ("(-) Influenciadores/Afiliados", "item", "afiliados", 1),
                ("= RESULTADO COMERCIAL", "subtotal", "resultado_comercial", 0),
            ],
        ),
        (
            "DESPESAS OPERACIONAIS",
            [
                ("(-) Consultoria", "item", None, 1),
                ("(-) Sistemas (Bling, ERP, Tiny, etc.)", "item", None, 1),
                ("(-) Salários", "item", None, 1),
                ("(-) Pró-labore", "item", None, 1),
                ("(-) Despesas Bancárias", "item", None, 1),
                ("(-) Contabilidade", "item", None, 1),
                ("= EBITDA (Resultado Operacional)", "subtotal", "ebitda", 0),
            ],
        ),
        (
            "RESULTADO FINANCEIRO",
            [
                ("(+) Outras Receitas Financeiras", "item", None, 1),
                ("(+) Rendimentos Financeiros", "item", None, 1),
                ("(-) Juros Bancários", "item", None, 1),
                ("(-) Multas", "item", None, 1),
                ("(-) IOF", "item", None, 1),
                ("(-) Tarifas Bancárias", "item", None, 1),
                ("= LUCRO LÍQUIDO", "total", "lucro_liquido", 0),
            ],
        ),
    ]

    manual_keys = {
        "(-) CMV (Custo das Mercadorias Vendidas)": "cmv",
        "(-) Custo com Expedição": "expedicao",
        "(-) Embalagens": "embalagens",
        "(-) Consultoria": "consultoria",
        "(-) Sistemas (Bling, ERP, Tiny, etc.)": "sistemas",
        "(-) Salários": "salarios",
        "(-) Pró-labore": "pro_labore",
        "(-) Despesas Bancárias": "despesas_bancarias",
        "(-) Contabilidade": "contabilidade",
        "(+) Outras Receitas Financeiras": "outras_receitas",
        "(+) Rendimentos Financeiros": "rendimentos",
        "(-) Juros Bancários": "juros",
        "(-) Multas": "multas",
        "(-) IOF": "iof",
        "(-) Tarifas Bancárias": "tarifas_bancarias",
    }

    current_row = header_row + 1
    key_to_row: dict[str, int] = {}
    row_meta: list[tuple[int, str, RowKind, str | None, int]] = []

    for section_title, lines in sections:
        ws.cell(row=current_row, column=1, value=section_title).font = LABEL_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        ws.cell(row=current_row, column=2).fill = SECTION_FILL
        current_row += 1

        for label, kind, key, indent in lines:
            display_label = ("  " * indent) + label
            label_cell = ws.cell(
                row=current_row, column=1, value=as_excel_label(display_label)
            )
            label_cell.border = BORDER
            if kind in ("subtotal", "total"):
                label_cell.font = LABEL_FONT
                label_cell.fill = TOTAL_FILL
            row_meta.append((current_row, label, kind, key, indent))
            if key:
                key_to_row[key] = current_row
            current_row += 1

    def ref(key: str) -> str:
        return _row_ref(key_to_row[key])

    data_formulas: dict[str, str] = {
        "faturamento_bruto": sum_columns(det_col, [SRC_FATURADO, SRC_FRETE], last_row),
        "vendas_canceladas": sum_cancelled_sales(det_col, SRC_FATURADO, last_row),
        "devolucoes": sum_return_cfops(det_col, SRC_FATURADO, last_row),
        "descontos": sum_column(det_col, SRC_DESCONTO, last_row),
        "impostos_vendas": f"={ref('receita_bruta_ajustada')}*{sales_tax_rate_cell}",
        "taxas_marketplace": sum_column(det_col, "CANDARU", last_row),
        "frete": sum_column(det_col, SRC_FRETE, last_row),
        "ads_ml": sumifs_wildcard(
            det_col, "ADS", "market_place", "*Mercado Livre*", last_row
        ),
        "ads_shopee": sumifs_wildcard(
            det_col, "ADS", "market_place", "*Shopee*", last_row
        ),
        "afiliados": sum_column(det_col, "Afiliado", last_row),
    }

    for row_num, label, kind, key, _indent in row_meta:
        value_cell = ws.cell(row=row_num, column=2)
        value_cell.border = BORDER
        value_cell.number_format = MONEY_FMT
        value_cell.alignment = Alignment(horizontal="right")

        stripped_label = label.strip()
        if stripped_label in manual_keys:
            key_to_row[manual_keys[stripped_label]] = row_num
            value_cell.value = 0
            continue

        if key and key in data_formulas:
            value_cell.value = data_formulas[key]
        elif kind in ("subtotal", "total"):
            value_cell.font = LABEL_FONT
            value_cell.fill = TOTAL_FILL

    ws[f"B{key_to_row['receita_bruta_ajustada']}"] = (
        f"={ref('faturamento_bruto')}"
        f"-{ref('vendas_canceladas')}"
        f"-{ref('devolucoes')}"
        f"-{ref('descontos')}"
    )
    ws[f"B{key_to_row['receita_liquida']}"] = (
        f"={ref('receita_bruta_ajustada')}-{ref('impostos_vendas')}"
    )
    ws[f"B{key_to_row['lucro_bruto']}"] = (
        f"={ref('receita_liquida')}"
        f"-B{key_to_row['cmv']}"
        f"-{ref('taxas_marketplace')}"
        f"-{ref('frete')}"
        f"-B{key_to_row['expedicao']}"
        f"-B{key_to_row['embalagens']}"
    )
    ws[f"B{key_to_row['resultado_comercial']}"] = (
        f"={ref('lucro_bruto')}-{ref('ads_ml')}-{ref('ads_shopee')}-{ref('afiliados')}"
    )
    ws[f"B{key_to_row['ebitda']}"] = (
        f"={ref('resultado_comercial')}"
        f"-B{key_to_row['consultoria']}"
        f"-B{key_to_row['sistemas']}"
        f"-B{key_to_row['salarios']}"
        f"-B{key_to_row['pro_labore']}"
        f"-B{key_to_row['despesas_bancarias']}"
        f"-B{key_to_row['contabilidade']}"
    )
    ws[f"B{key_to_row['lucro_liquido']}"] = (
        f"={ref('ebitda')}"
        f"+B{key_to_row['outras_receitas']}"
        f"+B{key_to_row['rendimentos']}"
        f"-B{key_to_row['juros']}"
        f"-B{key_to_row['multas']}"
        f"-B{key_to_row['iof']}"
        f"-B{key_to_row['tarifas_bancarias']}"
    )

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A4"

    note_row = current_row + 1
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "Linhas com valor 0 podem ser preenchidas manualmente (CMV, despesas operacionais, "
            "resultado financeiro). Demais valores são calculados a partir da aba Detalhado. "
            f"Impostos sobre vendas usam a taxa em {sales_tax_rate_cell}."
        ),
    )
    ws.cell(row=note_row, column=1).font = NOTE_FONT


def add_dre_to_workbook(
    workbook_path: Path, *, output_path: Path | None = None
) -> Path:
    """Add or replace the DRE sheet in an existing Fechamento workbook."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path)
    if "Detalhado" not in wb.sheetnames:
        raise ValueError(f"'Detalhado' sheet not found in {workbook_path}")

    ws_det = wb["Detalhado"]
    headers = [cell.value for cell in ws_det[1]]
    det_col = {
        str(name): get_column_letter(idx)
        for idx, name in enumerate(headers, start=1)
        if name
    }
    last_row = ws_det.max_row
    if last_row < 2:
        raise ValueError("Detalhado sheet has no data rows.")

    if "DRE" in wb.sheetnames:
        del wb["DRE"]
    ws_dre = wb.create_sheet("DRE")
    sales_tax_cell = "Parametros!$G$3" if "Parametros" in wb.sheetnames else "0.04"
    build_dre(ws_dre, det_col, last_row, sales_tax_rate_cell=sales_tax_cell)

    dest = output_path or workbook_path
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Existing relatorio_fechamento .xlsx (must contain Detalhado sheet)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output path (default: overwrite --input)",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if not args.input.is_file():
        print(f"Error: workbook not found: {args.input}")
        sys.exit(1)

    dest = add_dre_to_workbook(args.input, output_path=args.output)
    print(f"DRE sheet written to: {dest.resolve()}")


if __name__ == "__main__":
    main()
