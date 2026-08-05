"""Shared constants, styles, and helpers for Fechamento report workbooks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# Default assumption rates (editable in the Parametros sheet).
DEFAULT_ADS_RATES: dict[str, float] = {
    "Mercado Livre": 0.0597444848231328,
    "Mercado Livre Full": 0.0597444848231328,
    "Shopee": 0.0816910844699686,
    "Shopee Full": 0.0816910844699686,
    "TikTok": 0.0808623294375085,
}
DEFAULT_AFILIADO_RATES: dict[str, float] = {
    "Mercado Livre": 0.0102977242994446,
    "Mercado Livre Full": 0.0102977242994446,
    "Shopee": 0.031148329972608,
    "Shopee Full": 0.031148329972608,
    "TikTok": 0.0924554265462002,
}
DEFAULT_IRPJ_RATE = 0.0308
DEFAULT_SALES_TAX_RATE = 0.04
CANDARU_RATE_LOW = 0.03
CANDARU_RATE_DEFAULT = 0.07
CANDARU_LOW_MARKETPLACE_SUBSTRINGS = ("amazon b2b", "tiktok shop")

MONEY_FMT = "R$ #,##0.00"
PCT_FMT = "0.00%"
INT_FMT = "#,##0"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")
LABEL_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ADDED_COLUMNS = [
    "CFOP",
    "ADS",
    "Afiliado",
    "IRPJ/CSLL",
    "CANDARU",
]

SRC_MARKETPLACE = "market_place"
SRC_FATURADO = "valor_produtos"
SRC_BASE_COMISSAO = "valor_base_comissao"
SRC_BASE_ICMS = "valor_icms_bc"
SRC_DESCONTO = "valor_desconto"
SRC_FRETE = "valor_frete"
SRC_ICMS = "valor_icms"
SRC_DIFAL = "valor_icms_difal"
SRC_PIS = "valor_pis"
SRC_COFINS = "valor_cofins"
SRC_STATUS = "status"
SRC_NATUREZA = "natureza_operacao"

MONEY_COLUMNS = {
    "valor_produtos",
    "valor_frete",
    "valor_desconto",
    "valor_nf",
    "valor_base_comissao",
    "valor_icms",
    "valor_icms_difal",
    "valor_icms_bc",
    "valor_pis",
    "valor_cofins",
    "ADS",
    "Afiliado",
    "IRPJ/CSLL",
    "CANDARU",
}
TEXT_COLUMNS = {"chave_nfe", "destinatario_doc"}

DET_WIDTHS = {
    "market_place": 16,
    "chave_nfe": 36,
    "numero": 10,
    "serie": 7,
    "data_emissao": 22,
    "natureza_operacao": 20,
    "destinatario_doc": 16,
    "destinatario_uf": 8,
    "status": 22,
    "arquivo": 30,
}

# CFOPs comuns de devolução de venda (entrada de mercadoria devolvida pelo cliente).
RETURN_CFOPS = (
    "1201",
    "1202",
    "1203",
    "1204",
    "1410",
    "1411",
    "2201",
    "2202",
    "2203",
    "2204",
    "2410",
    "2411",
)

ITEM_HEADERS = [
    "market_place",
    "chave_nfe",
    "numero_nf",
    "data_emissao",
    "item",
    "codigo",
    "ean",
    "descricao",
    "ncm",
    "cfop",
    "unidade",
    "quantidade",
    "valor_unitario",
    "valor_total",
    "destinatario_doc",
    "destinatario_uf",
    "arquivo",
]

ITEMS_SHEET = "Items"
CUSTOS_SHEET = "Custos Produtos"
PRODUCT_COST_HEADERS = ["codigo", "ean", "descricao", "custo", "chave"]
ITEMS_ADDED_COLUMNS = ["custo_unitario", "custo_total"]

ITEM_MONEY_COLUMNS = {"valor_unitario", "valor_total", "custo_unitario", "custo_total"}
ITEM_TEXT_COLUMNS = {"chave_nfe", "codigo", "ean", "destinatario_doc"}

ITEM_WIDTHS = {
    "market_place": 16,
    "chave_nfe": 36,
    "numero_nf": 10,
    "data_emissao": 22,
    "codigo": 14,
    "ean": 16,
    "descricao": 36,
    "cfop": 8,
    "quantidade": 12,
    "custo_unitario": 14,
    "custo_total": 14,
}

CUSTOS_WIDTHS = {
    "codigo": 14,
    "ean": 16,
    "descricao": 36,
    "custo": 14,
    "chave": 24,
}


def candaru_rate_for(marketplace: str) -> float:
    name = marketplace.strip().casefold()
    if any(sub in name for sub in CANDARU_LOW_MARKETPLACE_SUBSTRINGS):
        return CANDARU_RATE_LOW
    return CANDARU_RATE_DEFAULT


def num(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_cell(value: object) -> str | int | float:
    if value is None:
        return ""
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def as_excel_label(value: str) -> str:
    """Return *value* safe for Excel label cells (avoids formula interpretation)."""
    stripped = value.lstrip()
    if stripped and stripped[0] in "=-+@":
        return f"'{value}"
    return value


def rows_as_dicts(ws: Worksheet) -> list[dict[str, object]]:
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    return [dict(zip(headers, row)) for row in rows]


def load_result(path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Invoices" not in wb.sheetnames:
        raise ValueError(f"'Invoices' sheet not found in {path}")
    invoices = rows_as_dicts(wb["Invoices"])
    items = rows_as_dicts(wb["Items"]) if "Items" in wb.sheetnames else []
    wb.close()
    return invoices, items


def cfop_by_invoice(items: list[dict[str, object]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in items:
        key = str(item.get("chave_nfe", ""))
        if key not in result and item.get("cfop"):
            result[key] = str(item.get("cfop"))
    return result


def product_key(codigo: object, ean: object) -> tuple[str, str]:
    return (str(codigo or "").strip(), str(ean or "").strip())


def deduplicate_products(items: list[dict[str, object]]) -> list[dict[str, str]]:
    """Return unique products keyed by (codigo, ean), preserving first description."""
    seen: dict[tuple[str, str], dict[str, str]] = {}
    for item in items:
        codigo, ean = product_key(item.get("codigo"), item.get("ean"))
        key = (codigo, ean)
        if key in seen:
            continue
        seen[key] = {
            "codigo": codigo,
            "ean": ean,
            "descricao": str(item.get("descricao", "")).strip(),
        }
    return [seen[key] for key in sorted(seen)]


def style_header_cell(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def det_range(det_col: dict[str, str], column: str, last_row: int) -> str:
    """Return an absolute column range on the Detalhado sheet (e.g. Detalhado!$E$2:$E$100)."""
    col = det_col[column]
    return f"Detalhado!${col}$2:${col}${last_row}"


def sum_column(det_col: dict[str, str], column: str, last_row: int) -> str:
    return f"=SUM({det_range(det_col, column, last_row)})"


def sum_columns(det_col: dict[str, str], columns: list[str], last_row: int) -> str:
    """Sum multiple Detalhado columns into one formula."""
    parts = [f"SUM({det_range(det_col, column, last_row)})" for column in columns]
    return f"={'+'.join(parts)}"


def sumifs_column(
    det_col: dict[str, str],
    sum_column_name: str,
    criteria_column: str,
    criteria: str,
    last_row: int,
) -> str:
    sum_rng = det_range(det_col, sum_column_name, last_row)
    crit_rng = det_range(det_col, criteria_column, last_row)
    return f'=SUMIFS({sum_rng},{crit_rng},"{criteria}")'


def sumifs_wildcard(
    det_col: dict[str, str],
    sum_column_name: str,
    criteria_column: str,
    pattern: str,
    last_row: int,
) -> str:
    sum_rng = det_range(det_col, sum_column_name, last_row)
    crit_rng = det_range(det_col, criteria_column, last_row)
    return f'=SUMIFS({sum_rng},{crit_rng},"{pattern}")'


def sum_return_cfops(det_col: dict[str, str], value_column: str, last_row: int) -> str:
    parts = [
        sumifs_column(det_col, value_column, "CFOP", cfop, last_row)
        for cfop in RETURN_CFOPS
    ]
    inner = "+".join(part.removeprefix("=") for part in parts)
    return f"={inner}"


def sum_cancelled_sales(
    det_col: dict[str, str], value_column: str, last_row: int
) -> str:
    sum_rng = det_range(det_col, value_column, last_row)
    status_rng = det_range(det_col, SRC_STATUS, last_row)
    natureza_rng = det_range(det_col, SRC_NATUREZA, last_row)
    return (
        f'=SUMIFS({sum_rng},{status_rng},"*Cancel*")'
        f'+SUMIFS({sum_rng},{natureza_rng},"*Cancel*")'
    )


def items_range(items_col: dict[str, str], column: str, last_row: int) -> str:
    col = items_col[column]
    return f"{ITEMS_SHEET}!${col}$2:${col}${last_row}"


def sum_items_column(items_col: dict[str, str], column: str, last_row: int) -> str:
    if last_row < 2:
        return "=0"
    return f"=SUM({items_range(items_col, column, last_row)})"
