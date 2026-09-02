"""Build a marketplace closing report (Fechamento) from the NF-e result workbook.

Reads the workbook produced by ``convert_nfe_xml_to_xlsx.py`` (sheets ``Invoices``
and ``Items``) and writes a new workbook with:

  - Resumo:     one clean grid, metrics as rows (incl. taxas, impostos, custos,
                lucro), marketplaces as columns + Total.
  - Detalhado:  every column from the Invoices sheet (kept verbatim) followed by the
                appended calculated columns (live Excel formulas).
  - Parametros: ADS, Afiliado, Custos Expedição and Frete Marketplace as period
                amounts (R$) entered manually per marketplace; CANDARU and global
                rates (IRPJ/CSLL, impostos sobre vendas) remain percentages.
                Resumo and DRE add Frete Marketplace into Taxas Marketplace.
  - Items:      line items from the NF-e workbook, with custo_unitario / custo_total.
  - Custos Produtos: deduplicated product list (codigo + ean) for manual cost entry.
  - DRE:        Demonstração do Resultado do Exercício (see ``build_dre.py``).
  - Conciliacao: how many invoices got an order number and a marketplace fee.

Only values that can be derived from the NF-e are included. KPIs that depend on
external data (Depositado, Custo Produto, Liquido, Margem, Frete Primario) are not
produced. The appended Detalhado columns are estimates computed from NF-e values:

  IRPJ/CSLL = taxa_irpj                   * Base ICMS (valor_icms_bc)
  CANDARU   = taxa_candaru(marketplace) * valor_base_comissao
              (3% when market_place contains AMAZON B2B or TIKTOK SHOP; 7% otherwise)

ADS, Afiliado, Custos Expedição and Frete Marketplace are marketplace totals
entered on Parametros (not allocated per NF-e). Amazon fees are 13.5% of
valor_base_comissao (3% for Amazon B2B) and do not use a marketplace spreadsheet.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from build_dre import build_dre
from build_items import build_custos_produtos, build_items
from marketplace_fees import (
    MARKETPLACE_FEE_COLUMN,
    FeeJoinStats,
    enrich_invoices_with_marketplace_fees,
)
from report_common import (
    ADDED_COLUMNS,
    BORDER,
    CANDARU_RATE_DEFAULT,
    CUSTOS_SHEET,
    DEFAULT_IRPJ_RATE,
    DEFAULT_SALES_TAX_RATE,
    DEFAULT_SPECIAL_ICMS_RATE,
    DET_WIDTHS,
    INT_FMT,
    ITEM_HEADERS,
    ITEMS_SHEET,
    LABEL_FONT,
    MONEY_COLUMNS,
    MONEY_FMT,
    NOTE_FONT,
    PCT_FMT,
    SRC_BASE_COMISSAO,
    SRC_BASE_ICMS,
    SRC_COFINS,
    SRC_DEST_UF,
    SRC_DIFAL,
    SRC_ICMS,
    SRC_MARKETPLACE,
    SRC_PIS,
    TEXT_COLUMNS,
    TITLE_FONT,
    TOTAL_FILL,
    as_cell,
    candaru_rate_for,
    cfop_by_invoice,
    deduplicate_products,
    load_result,
    num,
    style_header_cell,
)

# The Detalhado sheet keeps every column from the Invoices sheet (same names and
# order) and then appends the calculated columns below.


def build_parametros(
    ws: Worksheet,
    marketplaces: list[str],
    candaru_rates: dict[str, float],
    *,
    regime_especial: bool = False,
) -> dict[str, str]:
    ws["A1"] = "Parâmetros por Marketplace"
    ws["A1"].font = TITLE_FONT
    for col, name in zip(
        "ABCDEF",
        (
            "Marketplace",
            "ADS",
            "Afiliado",
            "Taxa CANDARU",
            "Custos Expedição",
            "Frete Marketplace",
        ),
    ):
        cell = ws[f"{col}2"]
        cell.value = name
        style_header_cell(cell)

    for idx, mp in enumerate(marketplaces, start=3):
        ws[f"A{idx}"] = mp
        ws[f"B{idx}"] = 0
        ws[f"C{idx}"] = 0
        ws[f"D{idx}"] = candaru_rates.get(mp, CANDARU_RATE_DEFAULT)
        ws[f"E{idx}"] = 0
        ws[f"F{idx}"] = 0
        ws[f"B{idx}"].number_format = MONEY_FMT
        ws[f"C{idx}"].number_format = MONEY_FMT
        ws[f"D{idx}"].number_format = PCT_FMT
        ws[f"E{idx}"].number_format = MONEY_FMT
        ws[f"F{idx}"].number_format = MONEY_FMT

    ws["H2"] = "Taxa IRPJ/CSLL"
    ws["H2"].font = LABEL_FONT
    ws["I2"] = DEFAULT_IRPJ_RATE
    ws["I2"].number_format = PCT_FMT

    ws["H3"] = "Taxa Impostos sobre Vendas"
    ws["H3"].font = LABEL_FONT
    ws["I3"] = DEFAULT_SALES_TAX_RATE
    ws["I3"].number_format = PCT_FMT

    ws["H4"] = "Regime Especial"
    ws["H4"].font = LABEL_FONT
    ws["I4"] = "ATIVO" if regime_especial else "INATIVO"
    ws["I4"].font = LABEL_FONT
    if regime_especial:
        ws["I4"].fill = TOTAL_FILL

    ws["H5"] = "Taxa ICMS Regime Especial"
    ws["H5"].font = LABEL_FONT
    ws["I5"] = DEFAULT_SPECIAL_ICMS_RATE
    ws["I5"].number_format = PCT_FMT

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 12
    ws.freeze_panes = "A3"
    note_row = 4 + len(marketplaces)
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "Preencha ADS, Afiliado, Custos Expedição e Frete Marketplace em R$ "
            "(valor do período). O Frete Marketplace compõe Taxas Marketplace. "
            "A aba Resumo busca esses valores automaticamente. "
            "Taxa CANDARU permanece em percentual. Quando o Regime Especial está "
            "ATIVO, vendas para fora de MG usam 1,3% da base de ICMS."
        ),
    )
    ws.cell(row=note_row, column=1).font = NOTE_FONT
    return {
        "irpj": "$I$2",
        "sales_tax": "$I$3",
        "special_tax_status": "$I$4",
        "special_icms_rate": "$I$5",
    }


def build_detalhado(
    ws: Worksheet,
    invoices: list[dict[str, object]],
    invoice_headers: list[str],
    invoice_cfops: dict[str, str],
    global_rate_cells: dict[str, str],
) -> tuple[dict[str, str], int]:
    """Write the Detalhado sheet and return (column_name -> letter, last_row).

    All Invoices columns are kept verbatim; the calculated columns are appended.
    """
    columns = list(invoice_headers) + ADDED_COLUMNS
    det_col = {name: get_column_letter(i) for i, name in enumerate(columns, start=1)}

    for col_idx, name in enumerate(columns, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx, value=name))

    irpj = f"Parametros!{global_rate_cells['irpj']}"
    regime_status = f"Parametros!{global_rate_cells['special_tax_status']}"
    special_icms_rate = f"Parametros!{global_rate_cells['special_icms_rate']}"

    ordered = sorted(invoices, key=lambda inv: num(inv.get("numero")))
    for offset, inv in enumerate(ordered):
        r = offset + 2
        chave = str(inv.get("chave_nfe", ""))

        for name in invoice_headers:
            ws[f"{det_col[name]}{r}"] = as_cell(inv.get(name, ""))

        ws[f"{det_col['CFOP']}{r}"] = invoice_cfops.get(chave, "")

        base_comissao = f"{det_col[SRC_BASE_COMISSAO]}{r}"
        base = f"{det_col[SRC_BASE_ICMS]}{r}"
        original_icms = f"{det_col[SRC_ICMS]}{r}"
        destination_uf = f"{det_col[SRC_DEST_UF]}{r}"
        mp = f"{det_col[SRC_MARKETPLACE]}{r}"
        icms_outside_mg = f"{base}*{special_icms_rate}"
        icms_by_uf = f'IF(UPPER(TRIM({destination_uf}))<>"MG",{icms_outside_mg},{original_icms})'
        ws[f"{det_col['ICMS Considerado']}{r}"] = (
            f'=IF({regime_status}="ATIVO",{icms_by_uf},{original_icms})'
        )
        ws[f"{det_col['IRPJ/CSLL']}{r}"] = f"={irpj}*{base}"
        ws[f"{det_col['CANDARU']}{r}"] = (
            f"=VLOOKUP({mp},Parametros!$A:$D,4,FALSE)*{base_comissao}"
        )

    last_row = len(invoices) + 1
    for name in columns:
        col = det_col[name]
        if name in MONEY_COLUMNS:
            for r in range(2, last_row + 1):
                ws[f"{col}{r}"].number_format = MONEY_FMT
        elif name in TEXT_COLUMNS:
            for r in range(2, last_row + 1):
                ws[f"{col}{r}"].number_format = "@"

    for name in columns:
        ws.column_dimensions[det_col[name]].width = DET_WIDTHS.get(name, 13)
    ws.freeze_panes = "A2"
    return det_col, last_row


def build_resumo(
    ws: Worksheet,
    marketplaces: list[str],
    det_col: dict[str, str],
    *,
    items_col: dict[str, str] | None = None,
    sales_tax_rate_cell: str = "Parametros!$I$3",
    regime_especial: bool = False,
) -> None:
    ws["A1"] = "Fechamento - Resumo por Marketplace"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Regime Especial: {'ATIVO' if regime_especial else 'INATIVO'}"
    ws["A2"].font = LABEL_FONT
    if regime_especial:
        ws["A2"].fill = TOTAL_FILL

    n_mp = len(marketplaces)
    label_col = 1
    first_mp_col = 2
    last_mp_col = first_mp_col + n_mp - 1
    total_col = last_mp_col + 1
    pct_col = total_col + 1

    header_row = 3
    style_header_cell(ws.cell(row=header_row, column=label_col, value="Metrica"))
    for i, mp in enumerate(marketplaces):
        style_header_cell(ws.cell(row=header_row, column=first_mp_col + i, value=mp))
    style_header_cell(ws.cell(row=header_row, column=total_col, value="Total"))
    style_header_cell(ws.cell(row=header_row, column=pct_col, value="% s/ Faturado"))

    det = "Detalhado"
    fcol = det_col[SRC_MARKETPLACE]
    items_mp = items_col["market_place"] if items_col else None
    items_custo = items_col["custo_total"] if items_col else None

    metrics: list[tuple[str, str, str | None, str]] = [
        ("Notas Fiscais", "count", None, INT_FMT),
        ("Faturado", "sum", det_col[SRC_BASE_COMISSAO], MONEY_FMT),
        (
            "Taxas Marketplace",
            "marketplace_fees",
            det_col[MARKETPLACE_FEE_COLUMN],
            MONEY_FMT,
        ),
        ("ICMS", "sum", det_col["ICMS Considerado"], MONEY_FMT),
        ("DIFAL", "sum", det_col[SRC_DIFAL], MONEY_FMT),
        ("PIS", "sum", det_col[SRC_PIS], MONEY_FMT),
        ("COFINS", "sum", det_col[SRC_COFINS], MONEY_FMT),
        ("IRPJ/CSLL", "sum", det_col["IRPJ/CSLL"], MONEY_FMT),
        ("ADS", "param", "2", MONEY_FMT),
        ("Afiliado", "param", "3", MONEY_FMT),
        ("CANDARU", "sum", det_col["CANDARU"], MONEY_FMT),
        (
            "Imposto sobre vendas (SIMPLES NACIONAL)",
            "sales_tax",
            None,
            MONEY_FMT,
        ),
        ("Custos Produtos", "items_sum", None, MONEY_FMT),
        ("Custos Expedição", "param", "5", MONEY_FMT),
        ("Lucro", "lucro", None, MONEY_FMT),
    ]

    row = header_row + 1
    row_of: dict[str, int] = {}
    for label, _kind, _metric_col, _fmt in metrics:
        row_of[label] = row
        row += 1

    faturado_row = row_of["Faturado"]
    last_data_row = row_of["Lucro"]

    def col_letter(col_idx: int) -> str:
        return get_column_letter(col_idx)

    for label, kind, metric_col, fmt in metrics:
        data_row = row_of[label]
        label_cell = ws.cell(row=data_row, column=label_col, value=label)
        label_cell.font = LABEL_FONT
        label_cell.border = BORDER
        if kind == "lucro":
            label_cell.fill = TOTAL_FILL

        for i, _mp in enumerate(marketplaces):
            c = first_mp_col + i
            mp_ref = f"{col_letter(c)}${header_row}"
            crit = f"{det}!${fcol}:${fcol}"
            if kind == "count":
                formula: str | int = f"=COUNTIF({crit},{mp_ref})"
            elif kind == "sum":
                rng = f"{det}!${metric_col}:${metric_col}"
                formula = f"=SUMIF({crit},{mp_ref},{rng})"
            elif kind == "marketplace_fees":
                rng = f"{det}!${metric_col}:${metric_col}"
                formula = (
                    f"=SUMIF({crit},{mp_ref},{rng})"
                    f"+IFERROR(VLOOKUP({mp_ref},Parametros!$A:$F,6,FALSE),0)"
                )
            elif kind == "items_sum":
                if items_mp and items_custo:
                    formula = (
                        f"=SUMIF({ITEMS_SHEET}!${items_mp}:${items_mp},"
                        f"Resumo!{mp_ref},{ITEMS_SHEET}!${items_custo}:${items_custo})"
                    )
                else:
                    formula = 0
            elif kind == "sales_tax":
                formula = f"={col_letter(c)}{faturado_row}*{sales_tax_rate_cell}"
            elif kind == "param":
                formula = (
                    f"=IFERROR(VLOOKUP({mp_ref},Parametros!$A:$F,{metric_col},FALSE),0)"
                )
            else:
                first_cost = f"{col_letter(c)}{row_of['Taxas Marketplace']}"
                last_cost = f"{col_letter(c)}{row_of['Custos Expedição']}"
                formula = f"={col_letter(c)}{faturado_row}-SUM({first_cost}:{last_cost})"

            cell = ws.cell(row=data_row, column=c, value=formula)
            cell.number_format = fmt
            cell.border = BORDER
            if kind == "lucro":
                cell.fill = TOTAL_FILL
                cell.font = LABEL_FONT

        first = f"{col_letter(first_mp_col)}{data_row}"
        last = f"{col_letter(last_mp_col)}{data_row}"
        total_cell = ws.cell(
            row=data_row, column=total_col, value=f"=SUM({first}:{last})"
        )
        total_cell.number_format = fmt
        total_cell.font = LABEL_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.border = BORDER

        pct_formula = None
        if kind != "count":
            fat_total = f"{col_letter(total_col)}{faturado_row}"
            pct_formula = (
                f"=IFERROR({col_letter(total_col)}{data_row}/{fat_total},0)"
            )
        pct_cell = ws.cell(row=data_row, column=pct_col, value=pct_formula)
        if pct_formula is not None:
            pct_cell.number_format = PCT_FMT
        pct_cell.border = BORDER
        if kind == "lucro":
            pct_cell.fill = TOTAL_FILL

    for i in range(n_mp + 3):
        col = col_letter(label_col + i)
        ws.column_dimensions[col].width = 42 if i == 0 else 15
    ws.freeze_panes = ws.cell(row=header_row + 1, column=first_mp_col).coordinate

    note_row = last_data_row + 2
    ws.cell(
        row=note_row,
        column=label_col,
        value=(
            "Taxas Marketplace inclui as taxas calculadas e o Frete Marketplace "
            "informado na aba Parametros. Para Amazon, a taxa calculada é 13,5% "
            "do valor da venda (3% no Amazon B2B). "
            f"Regime Especial está {'ATIVO' if regime_especial else 'INATIVO'}; "
            "quando ativo, o ICMS para fora de MG é 1,3% da base de ICMS. "
            f"Imposto sobre vendas = Faturado × {sales_tax_rate_cell}. "
            "Custos Produtos = SOMASE de Items!custo_total pelo marketplace do "
            "cabeçalho. Lucro = Faturado − soma das demais linhas de valor. "
            "ADS, Afiliado e Custos Expedição vêm da aba Parametros (valores em R$, "
            "preenchimento manual). IRPJ/CSLL e CANDARU são estimativas por taxa."
        ),
    )
    ws.cell(row=note_row, column=label_col).font = NOTE_FONT


def build_conciliacao(ws: Worksheet, stats: FeeJoinStats) -> None:
    """Write the reconciliation counts behind the marketplace fee join."""
    ws["A1"] = "Conciliação das Taxas de Marketplace"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    style_header_cell(ws.cell(row=header_row, column=1, value="Indicador"))
    style_header_cell(ws.cell(row=header_row, column=2, value="Valor"))

    lines: list[tuple[str, int | float, str]] = [
        ("Notas fiscais processadas", stats.invoices, INT_FMT),
        ("Notas com número de pedido", stats.orders_found, INT_FMT),
        ("Notas com taxa encontrada", stats.fees_found, INT_FMT),
        ("Notas sem número de pedido", stats.invoices_without_order, INT_FMT),
        (
            "Pedidos ausentes nas planilhas dos marketplaces",
            stats.orders_without_fee,
            INT_FMT,
        ),
        ("Total de taxas de marketplace", stats.fees_total, MONEY_FMT),
    ]

    row = header_row + 1
    for label, value, fmt in lines:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.border = BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.number_format = fmt
        value_cell.border = BORDER
        row += 1

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18

    note_row = row + 1
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "O número do pedido vem do relatório do Bling, exceto Shopee Full e "
            "Mercado Livre Full, que usam o xPed da NF-e. O pedido é procurado em "
            "todas as planilhas de marketplace informadas. Notas sem pedido e "
            "pedidos fora do período das planilhas ficam com taxa R$ 0."
        ),
    )
    ws.cell(row=note_row, column=1).font = NOTE_FONT


def build_report_from_data(
    invoices: list[dict[str, object]],
    items: list[dict[str, object]],
    output: Path | BinaryIO,
    *,
    bling_path: Path | None = None,
    marketplace_paths: dict[str, Path] | None = None,
    regime_especial: bool = False,
) -> dict[str, object]:
    if not invoices:
        raise ValueError("No invoices found in the input data.")

    invoices, fee_stats = enrich_invoices_with_marketplace_fees(
        invoices,
        bling_path=bling_path,
        marketplace_paths=marketplace_paths or {},
    )
    invoice_headers = [h for h in invoices[0].keys() if h]
    item_headers = [h for h in items[0].keys() if h] if items else list(ITEM_HEADERS)
    invoice_cfops = cfop_by_invoice(items)

    marketplaces = sorted(
        {
            str(inv.get("market_place", ""))
            for inv in invoices
            if inv.get("market_place")
        }
    )

    candaru_rates = {mp: candaru_rate_for(mp) for mp in marketplaces}

    wb = Workbook()
    ws_resumo = wb.active
    assert ws_resumo is not None
    ws_resumo.title = "Resumo"
    ws_det = wb.create_sheet("Detalhado")
    ws_custos = wb.create_sheet(CUSTOS_SHEET)
    ws_items = wb.create_sheet(ITEMS_SHEET)
    ws_par = wb.create_sheet("Parametros")
    ws_dre = wb.create_sheet("DRE")
    ws_conc = wb.create_sheet("Conciliacao")

    global_rate_cells = build_parametros(
        ws_par,
        marketplaces,
        candaru_rates,
        regime_especial=regime_especial,
    )
    det_col, last_row = build_detalhado(
        ws_det, invoices, invoice_headers, invoice_cfops, global_rate_cells
    )
    custos_col, _ = build_custos_produtos(ws_custos, items)
    items_col, items_last_row = build_items(ws_items, items, item_headers, custos_col)
    build_resumo(
        ws_resumo,
        marketplaces,
        det_col,
        items_col=items_col,
        sales_tax_rate_cell=f"Parametros!{global_rate_cells['sales_tax']}",
        regime_especial=regime_especial,
    )
    build_dre(
        ws_dre,
        det_col,
        last_row,
        sales_tax_rate_cell=f"Parametros!{global_rate_cells['sales_tax']}",
        items_col=items_col,
        items_last_row=items_last_row,
    )
    build_conciliacao(ws_conc, fee_stats)

    if isinstance(output, Path):
        output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "invoices": len(invoices),
        "items": len(items),
        "products": len(deduplicate_products(items)) if items else 0,
        "marketplaces": marketplaces,
        "fee_orders_found": fee_stats.orders_found,
        "fee_matches": fee_stats.fees_found,
        "regime_especial": regime_especial,
    }


def build_report(
    input_path: Path,
    output_path: Path,
    *,
    bling_path: Path | None = None,
    marketplace_paths: dict[str, Path] | None = None,
    regime_especial: bool = False,
) -> dict[str, object]:
    invoices, items = load_result(input_path)
    return build_report_from_data(
        invoices,
        items,
        output_path,
        bling_path=bling_path,
        marketplace_paths=marketplace_paths,
        regime_especial=regime_especial,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Result workbook from convert_nfe_xml_to_xlsx.py (Invoices/Items sheets)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output report path (default: exports/relatorio_fechamento_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--regime-especial",
        action="store_true",
        help="Use 1.3% of the ICMS base for sales outside MG.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if not args.input.is_file():
        print(f"Error: input workbook not found: {args.input}")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = (
        args.output or Path("exports") / f"relatorio_fechamento_{timestamp}.xlsx"
    )

    info = build_report(
        args.input,
        output_path,
        regime_especial=args.regime_especial,
    )

    marketplaces = info["marketplaces"]
    marketplaces_text = (
        ", ".join(marketplaces) if isinstance(marketplaces, list) else ""
    )
    print(f"Report built from {info['invoices']} invoices ({info['items']} items).")
    print(f"Products in cost sheet: {info.get('products', 0)}")
    print(f"Marketplaces: {marketplaces_text}")
    print(f"Report saved to: {output_path.resolve()}")


if __name__ == "__main__":
    main()
