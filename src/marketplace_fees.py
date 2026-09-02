"""Join NF-e invoices to marketplace exports and calculate marketplace fees."""

from __future__ import annotations

import re
import unicodedata
from collections.abc import Callable, Iterable, Iterator
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from xlrd import open_workbook

ORDER_NUMBER_COLUMN = "numero_pedido"
MARKETPLACE_FEE_COLUMN = "taxa_marketplace"
AMAZON_FEE_RATE = 0.135
AMAZON_B2B_FEE_RATE = 0.03

_BLING_INVOICE_ALIASES = ("numero", "numero da nota", "numero nf", "nota fiscal")
_BLING_ORDER_ALIASES = (
    "numero do pedido multiloja",
    "pedido multiloja",
    "numero do pedido",
)
_SHOPEE_TYPE_ALIASES = ("ver", "view", "tipo")
_SHOPEE_ORDER_ALIASES = (
    "id do pedido",
    "numero do pedido",
    "pedido",
    "order id",
)
_SHOPEE_SUBTOTAL_ALIASES = (
    "preco do produto",
    "subtotal de mercadoria",
    "subtotal de mercadorias",
    "product subtotal",
)
_SHOPEE_INCOME_ALIASES = (
    "quantia total lancada r",
    "quantia total lancada",
    "rendas do pedido",
    "renda do pedido",
    "order income",
    "order earnings",
)
_MERCADO_LIVRE_ORDER_ALIASES = (
    "numero da operacao",
    "n da operacao",
    "numero de venda",
    "numero da venda",
    "id da venda",
)
_MERCADO_LIVRE_FEE_ALIASES = (
    "valor total de tarifas desconto ja aplicado",
    "valor total de tarifas",
)
_MAGALU_ORDER_ALIASES = ("numero do pedido", "n do pedido", "pedido")
_MAGALU_SALE_ALIASES = (
    "valor total dos itens do pedido",
    "valor de venda",
)
_MAGALU_NET_ALIASES = (
    "valor liquido estimado a receber",
    "valor de venda comissao plataforma",
)
MAGALU_SHIPPING_FEE = 35.90
_TIKTOK_TYPE_ALIASES = ("tipo de transacao", "transaction type")
_TIKTOK_ORDER_ALIASES = (
    "id do pedido ajuste",
    "id do pedido",
    "order id",
)
_TIKTOK_SALES_ALIASES = (
    "vendas liquidas dos produtos",
    "vendas liquidas",
    "net product sales",
)
_TIKTOK_SETTLED_ALIASES = (
    "valor total a ser liquidado",
    "total settlement amount",
)
_BELEZA_NA_WEB_ORDER_ALIASES = ("numero do pedido", "n do pedido")
_BELEZA_NA_WEB_PRODUCTS_ALIASES = ("valor dos produtos",)
_BELEZA_NA_WEB_NET_ALIASES = ("valor repasse", "valor do repasse")


@dataclass(frozen=True)
class FeeJoinStats:
    """Matching totals from the two-stage invoice enrichment."""

    invoices: int
    orders_found: int
    fees_found: int
    orders_without_fee: int
    fees_total: float

    @property
    def invoices_without_order(self) -> int:
        return self.invoices - self.orders_found


def _repaired_text(text: str) -> str:
    """Undo UTF-8 bytes that were stored as Latin-1 code points."""
    try:
        return text.encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return text


def _normalized_text(value: object) -> str:
    text = _repaired_text(str(value or "").strip())
    text = unicodedata.normalize("NFKD", text.casefold())
    without_format = "".join(
        char
        for char in text
        if not unicodedata.combining(char) and unicodedata.category(char) != "Cf"
    )
    return " ".join(re.sub(r"[^a-z0-9]+", " ", without_format).split())


def _normalized_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", "", str(value).strip()).casefold()


def _normalized_invoice_number(value: object) -> str:
    number = _normalized_id(value)
    if number.isdigit():
        return number.lstrip("0") or "0"
    return number


def _money(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^\d,.\-]", "", text)
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        amount = float(text)
    except ValueError:
        return 0.0
    return -amount if negative else amount


def _workbook_rows(path: Path) -> Iterator[tuple[str, Iterable[tuple[object, ...]]]]:
    suffix = path.suffix.casefold()
    if suffix == ".xls":
        workbook = open_workbook(path, ignore_workbook_corruption=True)
        for sheet in workbook.sheets():
            yield sheet.name, (tuple(sheet.row_values(row)) for row in range(sheet.nrows))
        return
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Formato não suportado: {path.suffix}. Use .xls ou .xlsx.")

    # read_only=False: Shopee exports often set dimension to A1, which hides
    # the real header row (row 3 of the Renda sheet) in read-only mode.
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            yield worksheet.title, worksheet.iter_rows(values_only=True)
    finally:
        workbook.close()


def _column(row: tuple[object, ...], aliases: tuple[str, ...]) -> int | None:
    normalized_aliases = {_normalized_text(alias) for alias in aliases}
    for index, value in enumerate(row):
        if _normalized_text(value) in normalized_aliases:
            return index
    return None


def _required_columns(
    row: tuple[object, ...],
    definitions: dict[str, tuple[str, ...]],
) -> dict[str, int] | None:
    columns: dict[str, int] = {}
    for name, aliases in definitions.items():
        index = _column(row, aliases)
        if index is None:
            return None
        columns[name] = index
    return columns


def _value(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def load_bling_orders(path: Path) -> dict[str, str]:
    """Return normalized NF-e number -> normalized marketplace order number."""
    if not path.is_file():
        raise FileNotFoundError(f"Planilha do Bling não encontrada: {path}")

    definitions = {
        "invoice": _BLING_INVOICE_ALIASES,
        "order": _BLING_ORDER_ALIASES,
    }
    for _sheet_name, rows in _workbook_rows(path):
        iterator = iter(rows)
        for row in iterator:
            columns = _required_columns(row, definitions)
            if columns is None:
                continue

            orders: dict[str, str] = {}
            for data_row in iterator:
                invoice = _normalized_invoice_number(_value(data_row, columns["invoice"]))
                order = _normalized_id(_value(data_row, columns["order"]))
                if invoice and order:
                    orders[invoice] = order
            return orders

    raise ValueError(
        "A planilha do Bling não contém as colunas 'Número' e "
        "'Número do pedido multiloja'."
    )


def _sum_fees(
    rows: Iterable[tuple[object, ...]],
    columns: dict[str, int],
    *,
    left: str,
    right: str | None = None,
    filter_column: str | None = None,
    extra_per_order: float = 0.0,
) -> dict[str, float]:
    fees: dict[str, float] = {}
    for row in rows:
        if filter_column is not None:
            row_type = _normalized_text(_value(row, columns[filter_column]))
            if row_type not in {"order", "pedido"}:
                continue

        order = _normalized_id(_value(row, columns["order"]))
        if not order:
            continue
        if right is None:
            fee = _money(_value(row, columns[left]))
        else:
            fee = _money(_value(row, columns[left])) - _money(
                _value(row, columns[right])
            )
        fees[order] = round(fees.get(order, 0.0) + fee, 2)
    if extra_per_order:
        return {
            order: round(fee + extra_per_order, 2) for order, fee in fees.items()
        }
    return fees


def _load_fees(
    path: Path,
    *,
    marketplace: str,
    definitions: dict[str, tuple[str, ...]],
    expected_columns: str,
    left: str,
    right: str | None = None,
    filter_column: str | None = None,
    extra_per_order: float = 0.0,
) -> dict[str, float]:
    """Load fees from a known marketplace export after validating its columns."""
    if not path.is_file():
        raise FileNotFoundError(
            f"Planilha de taxas da {marketplace} não encontrada: {path}"
        )

    for _sheet_name, rows in _workbook_rows(path):
        iterator = iter(rows)
        for row in iterator:
            columns = _required_columns(row, definitions)
            if columns is not None:
                return _sum_fees(
                    iterator,
                    columns,
                    left=left,
                    right=right,
                    filter_column=filter_column,
                    extra_per_order=extra_per_order,
                )

    raise ValueError(
        f"A planilha '{path.name}' não é um relatório de taxas válido da "
        f"{marketplace}. São necessárias as colunas: {expected_columns}."
    )


def load_shopee_fees(path: Path) -> dict[str, float]:
    """Load Shopee fees by order from its Renda export."""
    return _load_fees(
        path,
        marketplace="Shopee",
        definitions={
            "type": _SHOPEE_TYPE_ALIASES,
            "order": _SHOPEE_ORDER_ALIASES,
            "subtotal": _SHOPEE_SUBTOTAL_ALIASES,
            "income": _SHOPEE_INCOME_ALIASES,
        },
        expected_columns=(
            "Ver, ID do pedido, Preço do produto e Quantia total lançada (R$)"
        ),
        left="subtotal",
        right="income",
        filter_column="type",
    )


def load_mercado_livre_fees(path: Path) -> dict[str, float]:
    """Load Mercado Livre fees by order from the sales reconciliation export.

    Taxa = Valor total de tarifas (desconto já aplicado). Linhas com o mesmo
    Número da operação são somadas.
    """
    return _load_fees(
        path,
        marketplace="Mercado Livre",
        definitions={
            "order": _MERCADO_LIVRE_ORDER_ALIASES,
            "fee": _MERCADO_LIVRE_FEE_ALIASES,
        },
        expected_columns=(
            "Número da operação e Valor total de tarifas (desconto já aplicado)"
        ),
        left="fee",
    )


def load_magalu_fees(path: Path) -> dict[str, float]:
    """Load Magalu fees by order from its sales export.

    Taxa = Valor total dos itens do pedido − Valor líquido estimado a
    receber + R$ 35,90 de envio, cobrado uma vez por pedido.
    """
    return _load_fees(
        path,
        marketplace="Magalu",
        definitions={
            "order": _MAGALU_ORDER_ALIASES,
            "sale": _MAGALU_SALE_ALIASES,
            "net": _MAGALU_NET_ALIASES,
        },
        expected_columns=(
            "Número do pedido, Valor total dos itens do pedido e Valor "
            "líquido estimado a receber"
        ),
        left="sale",
        right="net",
        extra_per_order=MAGALU_SHIPPING_FEE,
    )


def load_tiktok_fees(path: Path) -> dict[str, float]:
    """Load TikTok Shop fees by order from its income export.

    Taxa = Vendas líquidas dos produtos − Valor total a ser liquidado.
    """
    return _load_fees(
        path,
        marketplace="TikTok",
        definitions={
            "type": _TIKTOK_TYPE_ALIASES,
            "order": _TIKTOK_ORDER_ALIASES,
            "sales": _TIKTOK_SALES_ALIASES,
            "settled": _TIKTOK_SETTLED_ALIASES,
        },
        expected_columns=(
            "Tipo de transação, ID do pedido/ajuste, Vendas líquidas dos "
            "produtos e Valor total a ser liquidado"
        ),
        left="sales",
        right="settled",
        filter_column="type",
    )


def load_beleza_na_web_fees(path: Path) -> dict[str, float]:
    """Load Beleza na Web fees by order from its repasse export.

    Taxa = Valor dos produtos − Valor repasse.
    """
    return _load_fees(
        path,
        marketplace="Beleza na Web",
        definitions={
            "order": _BELEZA_NA_WEB_ORDER_ALIASES,
            "products": _BELEZA_NA_WEB_PRODUCTS_ALIASES,
            "net": _BELEZA_NA_WEB_NET_ALIASES,
        },
        expected_columns="Número do Pedido, Valor dos produtos e Valor repasse",
        left="products",
        right="net",
    )


@dataclass(frozen=True)
class MarketplaceFeeSource:
    """Fixed fee spreadsheet for one marketplace export format."""

    key: str
    label: str
    loader: Callable[[Path], dict[str, float]]
    help_text: str


FEE_SOURCES: tuple[MarketplaceFeeSource, ...] = (
    MarketplaceFeeSource(
        key="shopee",
        label="Shopee",
        loader=load_shopee_fees,
        help_text=(
            "Exportação Renda com as colunas Ver, ID do pedido, Preço do "
            "produto e Quantia total lançada (R$)."
        ),
    ),
    MarketplaceFeeSource(
        key="mercado_livre",
        label="Mercado Livre",
        loader=load_mercado_livre_fees,
        help_text=(
            "Relatório de conciliação por vendas com Número da operação e "
            "Valor total de tarifas (desconto já aplicado). Pedidos repetidos "
            "somam as tarifas."
        ),
    ),
    MarketplaceFeeSource(
        key="magalu",
        label="Magalu",
        loader=load_magalu_fees,
        help_text=(
            "Relatório de vendas/pedidos com Número do pedido, Valor total "
            "dos itens do pedido e Valor líquido estimado a receber. Taxa "
            "= itens − líquido + R$ 35,90 de envio por pedido."
        ),
    ),
    MarketplaceFeeSource(
        key="tiktok",
        label="TikTok",
        loader=load_tiktok_fees,
        help_text=(
            "Exportação de renda (aba Detalhes do pedido) com ID do "
            "pedido/ajuste, Vendas líquidas dos produtos e Valor total a "
            "ser liquidado. Taxa = vendas líquidas − valor liquidado."
        ),
    ),
    MarketplaceFeeSource(
        key="beleza_na_web",
        label="Beleza na Web",
        loader=load_beleza_na_web_fees,
        help_text=(
            "Planilha de repasse com Número do Pedido, Valor dos produtos e "
            "Valor repasse. Taxa = produtos − repasse."
        ),
    ),
)
FEE_SOURCE_BY_KEY = {source.key: source for source in FEE_SOURCES}


def _fees_for_paths(marketplace_paths: dict[str, Path]) -> dict[str, float]:
    sources: list[tuple[str, dict[str, float]]] = []
    for key, path in marketplace_paths.items():
        source = FEE_SOURCE_BY_KEY.get(key)
        if source is None:
            raise ValueError(f"Marketplace de taxas desconhecido: {key}")
        sources.append((source.label, source.loader(path)))
    return _merge_fee_maps(sources)


def _merge_fee_maps(
    sources: list[tuple[str, dict[str, float]]],
) -> dict[str, float]:
    merged: dict[str, float] = {}
    owners: dict[str, str] = {}
    for marketplace, fees in sources:
        for order, fee in fees.items():
            if order in owners:
                raise ValueError(
                    f"O pedido {order} aparece nas planilhas de "
                    f"{owners[order]} e {marketplace}; não é possível "
                    "determinar qual taxa usar."
                )
            owners[order] = marketplace
            merged[order] = fee
    return merged


def _amazon_fee(invoice: dict[str, object]) -> float:
    marketplace = _normalized_text(invoice.get("market_place"))
    rate = AMAZON_B2B_FEE_RATE if "b2b" in marketplace else AMAZON_FEE_RATE
    return round(_money(invoice.get("valor_base_comissao")) * rate, 2)


def _uses_xml_order(marketplace: object) -> bool:
    """Shopee Full and Mercado Livre Full store the order id in NF-e xPed."""
    name = _normalized_text(marketplace)
    if "full" not in name:
        return False
    return "shopee" in name or "mercado livre" in name


def _is_beleza_na_web(marketplace: object) -> bool:
    return "beleza na web" in _normalized_text(marketplace)


def _is_magalu(marketplace: object) -> bool:
    name = _normalized_text(marketplace)
    return "magalu" in name or "magazine luiza" in name


def _beleza_na_web_order_id(order: str) -> str:
    """Keep the Beleza na Web order number after the Bling hyphen prefix."""
    if "-" not in order:
        return order
    suffix = order.rsplit("-", 1)[-1]
    return suffix or order


def _magalu_order_id(order: str) -> str:
    """Magalu exports prefix orders with LU-; Bling usually omits it."""
    normalized = order.casefold()
    if normalized.startswith("lu-"):
        return normalized
    return f"lu-{normalized}"


def _order_for_invoice(
    invoice: dict[str, object],
    orders_by_invoice: dict[str, str],
) -> str:
    if _uses_xml_order(invoice.get("market_place")):
        return _normalized_id(invoice.get("xped"))
    invoice_number = _normalized_invoice_number(invoice.get("numero"))
    order = orders_by_invoice.get(invoice_number, "")
    if not order:
        return ""
    if _is_beleza_na_web(invoice.get("market_place")):
        return _beleza_na_web_order_id(order)
    if _is_magalu(invoice.get("market_place")):
        return _magalu_order_id(order)
    return order


def _fee_for_invoice(
    invoice: dict[str, object],
    *,
    order: str,
    fees_by_order: dict[str, float],
) -> tuple[float, bool]:
    """Return the invoice fee and whether a rate was applied."""
    if "amazon" in _normalized_text(invoice.get("market_place")):
        return _amazon_fee(invoice), True
    if order and order in fees_by_order:
        return fees_by_order[order], True
    return 0.0, False


def enrich_invoices_with_marketplace_fees(
    invoices: list[dict[str, object]],
    *,
    bling_path: Path | None,
    marketplace_paths: dict[str, Path],
) -> tuple[list[dict[str, object]], FeeJoinStats]:
    """Add order number and marketplace fee to invoice copies.

    Every order is looked up in each supplied marketplace source, regardless
    of the NF-e source folder. Amazon uses a fixed share of valor_base_comissao
    (3% for Amazon B2B, 13.5% otherwise) and does not need a marketplace
    spreadsheet. Shopee Full and Mercado Livre Full use xPed from the NF-e
    instead of the Bling report.
    """
    orders_by_invoice = load_bling_orders(bling_path) if bling_path else {}
    fees_by_order = _fees_for_paths(marketplace_paths)

    enriched: list[dict[str, object]] = []
    orders_found = 0
    fees_found = 0
    orders_without_fee = 0
    fees_total = 0.0
    for invoice in invoices:
        result = dict(invoice)
        order = _order_for_invoice(invoice, orders_by_invoice)
        fee, matched = _fee_for_invoice(
            invoice, order=order, fees_by_order=fees_by_order
        )

        if order:
            orders_found += 1
            if not matched:
                orders_without_fee += 1
        if matched:
            fees_found += 1
            fees_total += fee

        result[ORDER_NUMBER_COLUMN] = order
        result[MARKETPLACE_FEE_COLUMN] = fee
        enriched.append(result)

    return enriched, FeeJoinStats(
        invoices=len(invoices),
        orders_found=orders_found,
        fees_found=fees_found,
        orders_without_fee=orders_without_fee,
        fees_total=round(fees_total, 2),
    )
