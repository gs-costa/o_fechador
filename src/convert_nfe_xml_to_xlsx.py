"""Convert NF-e XML invoices from marketplace folders to XLSX.

Expects a root folder (default: NFs/) with one subfolder per marketplace,
each containing .xml files. The subfolder name is written to the market_place column.

By default writes one consolidated workbook with:
  - Invoices: one row per NF-e
  - Items: one row per product line

After the database is written, the Fechamento report is built automatically from it
(use --no-report to skip, --cost-base to include product costs).

Use --per-file to also write one .xlsx next to each .xml.
"""

from __future__ import annotations

import argparse
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# build_report lives next to this script; ensure it is importable when run directly.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_report import build_report
from report_common import ITEM_HEADERS

NFE_NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

INVOICE_HEADERS = [
    "market_place",
    "chave_nfe",
    "numero",
    "serie",
    "data_emissao",
    "natureza_operacao",
    "destinatario_doc",
    "destinatario_uf",
    "valor_produtos",
    "valor_frete",
    "valor_desconto",
    "valor_nf",
    "valor_base_comissao",
    "valor_icms",
    "valor_icms_difal",
    "valor_icms_bc",
    "valor_pis",
    "valor_cofins",
    "status",
    "arquivo",
]

def _text(parent: ET.Element | None, path: str, default: str = "") -> str:
    if parent is None:
        return default
    node = parent.find(path, NFE_NS)
    if node is None or node.text is None:
        return default
    return node.text.strip()


def _doc_from_person(person: ET.Element | None) -> str:
    if person is None:
        return ""
    return _text(person, "nfe:CNPJ") or _text(person, "nfe:CPF")


def parse_nfe(
    xml_path: Path,
    *,
    market_place: str,
) -> tuple[dict[str, str | float], list[dict[str, str | float]]]:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    inf_nfe = root.find(".//nfe:infNFe", NFE_NS)
    if inf_nfe is None:
        raise ValueError("infNFe not found — not a valid NF-e XML")

    ide = inf_nfe.find("nfe:ide", NFE_NS)
    dest = inf_nfe.find("nfe:dest", NFE_NS)
    total = inf_nfe.find("nfe:total/nfe:ICMSTot", NFE_NS)
    prot = root.find(".//nfe:infProt", NFE_NS)

    chave = (inf_nfe.get("Id") or "").removeprefix("NFe") or _text(prot, "nfe:chNFe")
    ender_dest = dest.find("nfe:enderDest", NFE_NS) if dest is not None else None

    valor_frete = float(_text(total, "nfe:vFrete", "0") or 0)
    valor_nf = float(_text(total, "nfe:vNF", "0") or 0)

    invoice: dict[str, str | float] = {
        "market_place": market_place,
        "chave_nfe": chave,
        "numero": _text(ide, "nfe:nNF"),
        "serie": _text(ide, "nfe:serie"),
        "data_emissao": _text(ide, "nfe:dhEmi"),
        "natureza_operacao": _text(ide, "nfe:natOp"),
        # "emitente_cnpj": _text(emit, "nfe:CNPJ"),
        # "emitente_nome": _text(emit, "nfe:xNome"),
        # "emitente_fantasia": _text(emit, "nfe:xFant"),
        "destinatario_doc": _doc_from_person(dest),
        # "destinatario_nome": _text(dest, "nfe:xNome"),
        "destinatario_uf": _text(ender_dest, "nfe:UF"),
        # "destinatario_municipio": _text(ender_dest, "nfe:xMun"),
        "valor_produtos": float(_text(total, "nfe:vProd", "0") or 0),
        "valor_frete": valor_frete,
        "valor_desconto": float(_text(total, "nfe:vDesc", "0") or 0),
        "valor_nf": valor_nf,
        "valor_base_comissao": valor_nf - valor_frete,
        "valor_icms": float(_text(total, "nfe:vICMS", "0") or 0),
        "valor_icms_difal": float(_text(total, "nfe:vICMSUFDest", "0") or 0),
        "valor_icms_bc": float(_text(total, "nfe:vBC", "0") or 0),
        "valor_pis": float(_text(total, "nfe:vPIS", "0") or 0),
        "valor_cofins": float(_text(total, "nfe:vCOFINS", "0") or 0),
        # "valor_tot_trib": float(_text(total, "nfe:vTotTrib", "0") or 0),
        # "protocolo": _text(prot, "nfe:nProt"),
        "status": _text(prot, "nfe:xMotivo"),
        "arquivo": xml_path.name,
    }

    items: list[dict[str, str | float]] = []
    for det in inf_nfe.findall("nfe:det", NFE_NS):
        prod = det.find("nfe:prod", NFE_NS)
        items.append(
            {
                "market_place": market_place,
                "chave_nfe": chave,
                "numero_nf": invoice["numero"],
                "data_emissao": invoice["data_emissao"],
                "item": det.get("nItem", ""),
                "codigo": _text(prod, "nfe:cProd"),
                "ean": _text(prod, "nfe:cEAN"),
                "descricao": _text(prod, "nfe:xProd"),
                "ncm": _text(prod, "nfe:NCM"),
                "cfop": _text(prod, "nfe:CFOP"),
                "unidade": _text(prod, "nfe:uCom"),
                "quantidade": float(_text(prod, "nfe:qCom", "0") or 0),
                "valor_unitario": float(_text(prod, "nfe:vUnCom", "0") or 0),
                "valor_total": float(_text(prod, "nfe:vProd", "0") or 0),
                "destinatario_doc": invoice["destinatario_doc"],
                "destinatario_uf": invoice["destinatario_uf"],
                "arquivo": xml_path.name,
            }
        )

    return invoice, items


def _write_sheet(ws, headers: list[str], rows: list[dict[str, str | float]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws.iter_cols(
            min_col=col_idx, max_col=col_idx, min_row=2, values_only=True
        ):
            for value in cell:
                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def write_consolidated_xlsx(
    invoices: list[dict[str, str | float]],
    items: list[dict[str, str | float]],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws_inv = wb.active
    ws_inv.title = "Invoices"
    _write_sheet(ws_inv, INVOICE_HEADERS, invoices)

    ws_items = wb.create_sheet("Items")
    _write_sheet(ws_items, ITEM_HEADERS, items)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_single_xlsx(
    invoice: dict[str, str | float],
    items: list[dict[str, str | float]],
    output_path: Path,
) -> None:
    write_consolidated_xlsx([invoice], items, output_path)


def discover_marketplace_folders(root: Path) -> list[tuple[str, Path]]:
    if not root.is_dir():
        raise FileNotFoundError(f"Folder not found: {root}")

    if list(root.glob("*.xml")):
        return [(root.name, root)]

    marketplaces = [
        (subdir.name, subdir)
        for subdir in sorted(root.iterdir())
        if subdir.is_dir() and list(subdir.glob("*.xml"))
    ]
    if not marketplaces:
        raise FileNotFoundError(f"No .xml files found in {root} or its subfolders")
    return marketplaces


def convert_marketplaces(
    root_dir: Path,
    output_path: Path,
    *,
    per_file: bool = False,
) -> tuple[int, int, list[str], list[str]]:
    marketplaces = discover_marketplace_folders(root_dir)

    all_invoices: list[dict[str, str | float]] = []
    all_items: list[dict[str, str | float]] = []
    all_errors: list[str] = []
    processed: list[str] = []

    for marketplace_name, marketplace_dir in marketplaces:
        invoices: list[dict[str, str | float]] = []
        items: list[dict[str, str | float]] = []
        errors: list[str] = []

        for xml_path in sorted(marketplace_dir.glob("*.xml")):
            try:
                invoice, invoice_items = parse_nfe(
                    xml_path, market_place=marketplace_name
                )
            except Exception as exc:  # noqa: BLE001 - collect and continue
                errors.append(f"{marketplace_name}/{xml_path.name}: {exc}")
                continue

            invoices.append(invoice)
            items.extend(invoice_items)

            if per_file:
                write_single_xlsx(
                    invoice, invoice_items, marketplace_dir / f"{xml_path.stem}.xlsx"
                )

        all_invoices.extend(invoices)
        all_items.extend(items)
        all_errors.extend(errors)
        if invoices:
            processed.append(f"{marketplace_name} ({len(invoices)} invoices)")

    if not all_invoices:
        raise FileNotFoundError(f"No invoices converted from {root_dir}")

    write_consolidated_xlsx(all_invoices, all_items, output_path)
    return len(all_invoices), len(all_items), all_errors, processed


def build_parser() -> argparse.ArgumentParser:
    default_input = Path("NFs")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=default_input,
        help="Root folder with marketplace subfolders, or a single marketplace folder",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Path for the consolidated .xlsx (default: exports/nfe_invoices_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--per-file",
        action="store_true",
        help="Also write one .xlsx per .xml",
    )
    parser.add_argument(
        "--no-report",
        action="store_true",
        help="Skip building the Fechamento report after conversion",
    )
    parser.add_argument(
        "--report-output",
        type=Path,
        default=None,
        help="Report path (default: exports/relatorio_fechamento_<timestamp>.xlsx)",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    input_dir: Path = args.input_dir

    if not input_dir.is_dir():
        print(f"Error: folder not found: {input_dir}")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or Path("exports") / f"nfe_invoices_{timestamp}.xlsx"

    print(f"Reading XML files from: {input_dir.resolve()}")
    try:
        n_invoices, n_items, errors, processed = convert_marketplaces(
            input_dir,
            output_path,
            per_file=args.per_file,
        )
    except FileNotFoundError as exc:
        print(f"Error: {exc}")
        sys.exit(1)

    print(f"Converted {n_invoices} invoices ({n_items} line items).")
    for marketplace_summary in processed:
        print(f"  - {marketplace_summary}")
    print(f"Consolidated XLSX: {output_path.resolve()}")
    if args.per_file:
        print("Per-file XLSX written next to each .xml in its marketplace folder.")
    if errors:
        print(f"Skipped {len(errors)} file(s):")
        for err in errors[:20]:
            print(f"  - {err}")
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more")

    if args.no_report:
        return

    report_output = (
        args.report_output or Path("exports") / f"relatorio_fechamento_{timestamp}.xlsx"
    )
    print("\nBuilding Fechamento report...")
    try:
        build_report(output_path, report_output)
    except Exception as exc:  # noqa: BLE001 - report is a best-effort follow-up step
        print(f"Warning: could not build report: {exc}")
        return

    print(f"Report saved to: {report_output.resolve()}")


if __name__ == "__main__":
    main()
