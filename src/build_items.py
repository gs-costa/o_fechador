"""Build Items and Custos Produtos sheets for the Fechamento report.

Copies line items from the NF-e workbook, creates a deduplicated product-cost
lookup sheet (keyed by codigo + ean), and appends CMV-related calculated columns.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from report_common import (
    CUSTOS_SHEET,
    CUSTOS_WIDTHS,
    ITEM_MONEY_COLUMNS,
    ITEM_TEXT_COLUMNS,
    ITEM_WIDTHS,
    ITEMS_ADDED_COLUMNS,
    MONEY_FMT,
    PRODUCT_COST_HEADERS,
    as_cell,
    deduplicate_products,
    num,
    style_header_cell,
)

SRC_CODIGO = "codigo"
SRC_EAN = "ean"
SRC_QUANTIDADE = "quantidade"


def build_custos_produtos(
    ws: Worksheet,
    items: list[dict[str, object]],
) -> tuple[dict[str, str], int]:
    """Write the Custos Produtos sheet and return (column_name -> letter, last_row)."""
    products = deduplicate_products(items)
    col = {
        name: get_column_letter(idx) for idx, name in enumerate(PRODUCT_COST_HEADERS, start=1)
    }

    for col_idx, name in enumerate(PRODUCT_COST_HEADERS, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx, value=name))

    for offset, product in enumerate(products):
        row = offset + 2
        ws[f"{col['codigo']}{row}"] = product["codigo"]
        ws[f"{col['ean']}{row}"] = product["ean"]
        ws[f"{col['descricao']}{row}"] = product["descricao"]
        ws[f"{col['custo']}{row}"] = ""
        ws[f"{col['chave']}{row}"] = f"={col['codigo']}{row}&\"|\"&{col['ean']}{row}"

        ws[f"{col['codigo']}{row}"].number_format = "@"
        ws[f"{col['ean']}{row}"].number_format = "@"
        ws[f"{col['custo']}{row}"].number_format = MONEY_FMT

    last_row = len(products) + 1
    for name, width in CUSTOS_WIDTHS.items():
        ws.column_dimensions[col[name]].width = width
    ws.column_dimensions[col["chave"]].hidden = True
    ws.freeze_panes = "A2"
    return col, last_row


def _custo_unitario_formula(
    row: int,
    items_col: dict[str, str],
    custos_col: dict[str, str],
) -> str:
    codigo = f"${items_col[SRC_CODIGO]}${row}"
    ean = f"${items_col[SRC_EAN]}${row}"
    chave_col = custos_col["chave"]
    custo_col = custos_col["custo"]
    return (
        f"=IFERROR(INDEX('{CUSTOS_SHEET}'!${custo_col}:${custo_col},"
        f"MATCH({codigo}&\"|\"&{ean},'{CUSTOS_SHEET}'!${chave_col}:${chave_col},0)),0)"
    )


def build_items(
    ws: Worksheet,
    items: list[dict[str, object]],
    item_headers: list[str],
    custos_col: dict[str, str],
) -> tuple[dict[str, str], int]:
    """Write the Items sheet and return (column_name -> letter, last_row)."""
    columns = list(item_headers) + ITEMS_ADDED_COLUMNS
    items_col = {name: get_column_letter(idx) for idx, name in enumerate(columns, start=1)}

    for col_idx, name in enumerate(columns, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx, value=name))

    ordered = sorted(items, key=lambda item: (str(item.get("chave_nfe", "")), num(item.get("item"))))
    for offset, item in enumerate(ordered):
        row = offset + 2
        for name in item_headers:
            ws[f"{items_col[name]}{row}"] = as_cell(item.get(name, ""))

        custo_unit = _custo_unitario_formula(row, items_col, custos_col)
        ws[f"{items_col['custo_unitario']}{row}"] = custo_unit
        qty = f"{items_col[SRC_QUANTIDADE]}{row}"
        ws[f"{items_col['custo_total']}{row}"] = f"={qty}*{items_col['custo_unitario']}{row}"

    last_row = len(items) + 1
    for name in columns:
        col = items_col[name]
        if name in ITEM_MONEY_COLUMNS:
            for row in range(2, last_row + 1):
                ws[f"{col}{row}"].number_format = MONEY_FMT
        elif name in ITEM_TEXT_COLUMNS:
            for row in range(2, last_row + 1):
                ws[f"{col}{row}"].number_format = "@"

    for name in columns:
        ws.column_dimensions[items_col[name]].width = ITEM_WIDTHS.get(name, 13)
    ws.freeze_panes = "A2"
    return items_col, last_row
